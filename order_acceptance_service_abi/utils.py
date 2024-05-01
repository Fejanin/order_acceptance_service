from openpyxl import load_workbook


XLSXCOL_TO_MODELCOL = {
    'Код единицы продаж': 'sales_unit_code',
    'Код продукта': 'product_code',
    'Номер варианта': 'option_number',
    'Штрих-код ': 'barcode',
    'Вес нетто штуки, кг': 'net_weight_piece',
    'Кол-во штук в коробе, шт': 'number_pieces_in_box',
    'Вес нетто короба, кг': 'net_weight_of_box',
    'Вес брутто короба, кг': 'gross_weight_of_box',
    'Кол-во кор. на паллте, шт': 'number_of_boxes_per_pallet',
    'Коробок в слое': 'boxes_in_layer',
    'Завод': 'factory',
    'Срок годности, сут.': 'expiration_date',
    'Наименование': 'product_name',
    'Ед. изм.': 'units_measurement',
    }


def products_by_cat_dict(products):
    products_by_cat = {}
    for i in products:
        if not i.brand_name in products_by_cat:
            products_by_cat[i.brand_name] = {
                i.category: {
                    i.product_group: [i]
                }
            }
        else:
            if not i.category in products_by_cat[i.brand_name]:
                products_by_cat[i.brand_name][i.category] = {i.product_group: [i]}
            else:
                if not i.product_group in products_by_cat[i.brand_name][i.category]:
                    products_by_cat[i.brand_name][i.category][i.product_group] = [i]
                else:
                    products_by_cat[i.brand_name][i.category][i.product_group].append(i)
    return products_by_cat


def products_table_lst(prod_dict):
    result = []
    for brand_name in prod_dict:
        result.append({'brand_name': brand_name})
        for category in prod_dict[brand_name]:
            result.append({'category': category})
            for product_group in prod_dict[brand_name][category]:
                result.append({'product_group': product_group})
                result.append(prod_dict[brand_name][category][product_group])
    return result


def handle_uploaded_file(f):
    with open(f'uploads/{f.name}', 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)


def upgrade_products_table(file_name):
    '''
    Получает данные из xlsx-файла и преобразовывает их в список удобный для переноса в БД
    :param file_name: имя файла
    :return: список продуктов для таблицы Product
    '''
    wb = load_workbook(f'uploads/{file_name}', data_only=True)
    ws = wb.active
    rows = ws.max_row
    cols = ws.max_column
    keys_sku = {
        'Код единицы продаж': None,
        'Код продукта': None,
        'Номер варианта': None,
        'Штрих-код ': None,
        'Вес нетто штуки, кг': None,
        'Кол-во штук в коробе, шт': None,
        'Вес нетто короба, кг': None,
        'Вес брутто короба, кг': None,
        'Кол-во кор. на паллте, шт': None,
        'Коробок в слое': None,
        'Завод': None,
        'Срок годности, сут.': None,
        'Наименование': None,
        'Ед. изм.': None,
    }
    for c in range(1, cols + 1):
        cell = ws.cell(row=1, column=c).value
        if cell in keys_sku:
            keys_sku[cell] = c
    keys_levels = {
        (16.0, 'FF651C32', '00000000'): 'brand_name',
        (11.0, 'FF651C32', 'FFF2DEA6'): 'category',
        (11.0, 'standart', 'FFF2DEA6'): 'product_group',
    }
    current_levels = {
        'brand_name': None,
        'category': None,
        'product_group': None,
    }
    result = []
    for i in range(2, rows + 1):
        data_sku = {}
        if not ws.cell(row=i, column=1).value:
            continue
        cell = ws.cell(row=i, column=1)
        try:
            font_color = cell.font.color.rgb
        except:
            font_color = 'standart'
        font_size = cell.font.size
        fill = cell.fill.fgColor.rgb
        current_set = (font_size, font_color, fill)
        if current_set in keys_levels:
            current_levels[keys_levels[current_set]] = cell.value
            continue
        if 'SU' in cell.value:
            for j in keys_sku:
                cell_value = ws.cell(row=i, column=keys_sku[j]).value
                if j in XLSXCOL_TO_MODELCOL:
                    data_sku[XLSXCOL_TO_MODELCOL[j]] = cell_value
                else:
                    data_sku[j] = cell_value
            data_sku.update(current_levels)
            result.append(data_sku)
    return result
